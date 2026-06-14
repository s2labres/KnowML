
import openpyxl
import xlwings

class TestUtil: 
    """
    This class include utility functions for running tests.
    """

    @staticmethod
    def read_excel_sheet_to_arrays(file_path :str, packet_features: list, connection_features: list):
        """
        Read an excel workbook with sheets to an array of arrays.
        Assumes that first row is the header and the first column is the index.
        The index column in sheets is used to determine position in the outer array. 
        The COLUMN_ORDER is a dictionary that maps the column name to the position in the inner array.

        :param file_path (str): The path to the excel file.
        :return: tuple: A tuple of two arrays, where the first array is the packet data and the second array is the connection data and the third array is the destination mac port netstat data.
        """

        # Cache the evaluated excel file to ensure that the formulas are evaluated
        # TestUtil.evaluate_excel(file_path)

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        workbook_arr = []

        for sheet in workbook.worksheets:
            column_order = TestUtil.list_to_dic(packet_features)
            if sheet.title != "mocked packets":
                column_order = TestUtil.list_to_dic(connection_features)
            headers = list(next(sheet.iter_rows(max_row=1, values_only=True)))
            index_column = 0  # Assume the first column in the header is index

            data = [] # Store the data in the sheet
            position_arr = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[index_column] is None: # Skip empty rows
                    continue
                
                position_arr.append(row[index_column]) # Store the index in the position array
                row_data = [None] * len(column_order)
                for i, cell in enumerate(row[1:]):
                    
                    column_name = headers[i + 1] # Skip the index column
                    # Skip empty columns 
                    if column_name is not None and column_name != "num_connection_per_domain":
                        # Order according to the column_order
                        row_data[column_order[column_name]] = cell
                data.append(row_data)

            # Assume that None is empty string
            data = [[cell if cell is not None else "" for cell in row] for row in data]
            sorted_data = TestUtil.reorder_array(data, position_arr)

            workbook_arr.append(sorted_data)

        return workbook_arr[0], workbook_arr[1], workbook_arr[2]
    
    @staticmethod
    def reorder_array(array: list, position_array: list): 
        """
        Reorder an array based on the position array.
        Assumes that the position array has the same length as the array.

        Parameters:
            array (list): The array to be reordered.
            position_array (list): The array that contains the new position of the elements in the array.
        """

        sorted_data = [None] * len(array)
        for i in range(len(array)): 
                sorted_data[position_array[i]] = array[i]

        return sorted_data

    @staticmethod
    def list_to_dic(arr: list[str])->dict:
        """
        Convert a list to a dictionary, where the key is the element and the value is the index.
        :param arr (list): The list to convert.
        :return: dict: The dictionary representation of the list.
        """
        return {value: index for index, value in enumerate(arr)}
    
    @staticmethod
    def evaluate_excel(path: str)->None: 
        """
        Evaluate the formulas in an excel sheet and save.

        :param path (str): The path to the excel file.

        Note: This function MUST be called if the excel file has formulas and file wasn't cached. 
        Ref: https://stackoverflow.com/questions/36116162/python-openpyxl-data-only-true-returning-none/72901927#72901927
        """
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(path)
        excel_book.save()
        excel_book.close()
        excel_app.quit()